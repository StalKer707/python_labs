import csv
from pathlib import Path
from typing import Union

try:
    from openpyxl import Workbook
    from openpyxl.utils import get_column_letter
except ImportError:
    raise ImportError(
        "Для работы с XLSX требуется библиотека openpyxl. "
        "Установите её: pip install openpyxl"
    )


def csv_to_xlsx(csv_path: Union[str, Path], xlsx_path: Union[str, Path]) -> None:
    csv_p = Path(csv_path)
    xlsx_p = Path(xlsx_path)
    if not csv_p.exists():
        raise FileNotFoundError(f"CSV файл не найден: {csv_path}")
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    row_count = 0
    max_widths = {}
    with csv_p.open(encoding="utf-8") as f:
        reader = csv.reader(f)
        for row_idx, row in enumerate(reader, start=1):
            if not row:
                continue
            ws.append(row)
            row_count += 1
            for col_idx, cell_value in enumerate(row, start=1):
                current_width = len(str(cell_value))
                if col_idx not in max_widths:
                    max_widths[col_idx] = 0
                max_widths[col_idx] = max(max_widths[col_idx], current_width)
    if row_count == 0:
        raise ValueError("Пустой CSV файл")
    for col_idx, width in max_widths.items():
        column_letter = get_column_letter(col_idx)
        adjusted_width = max(width + 2, 8)
        ws.column_dimensions[column_letter].width = adjusted_width
    xlsx_p.parent.mkdir(parents=True, exist_ok=True)
    wb.save(xlsx_p)
    print(f"Успешно: Файл Excel сoхранен")